from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from delivery_vlm.delivery_schema import (
    XLSX_ORIGINAL_IMAGE_EMBED_COLUMN,
    XLSX_ORIGINAL_IMAGE_PATH_COLUMN,
)

_log = logging.getLogger(__name__)

_THUMB_MAX_PX = 120


def _write_sheet(
    wb: Workbook,
    *,
    title: str,
    rows: list[dict[str, Any]],
    columns: Sequence[str],
    embed_original_thumbnail: bool = False,
) -> None:
    ws = wb.create_sheet(title=title)
    cols = tuple(columns)
    path_i: int | None = None
    img_i: int | None = None
    if embed_original_thumbnail:
        try:
            path_i = cols.index(XLSX_ORIGINAL_IMAGE_PATH_COLUMN) + 1
            img_i = cols.index(XLSX_ORIGINAL_IMAGE_EMBED_COLUMN) + 1
        except ValueError:
            path_i = img_i = None

    for c, h in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(cols, start=1):
            if embed_original_thumbnail and key == XLSX_ORIGINAL_IMAGE_EMBED_COLUMN:
                ws.cell(row=r, column=c, value="")
                continue
            v = row.get(key, "")
            if v is None:
                v = ""
            ws.cell(row=r, column=c, value=v)

    if embed_original_thumbnail and path_i is not None and img_i is not None:
        for r, row in enumerate(rows, start=2):
            pth = str(row.get(XLSX_ORIGINAL_IMAGE_PATH_COLUMN, "")).strip()
            if not pth:
                continue
            fp = Path(pth)
            if not fp.is_file():
                continue
            try:
                im = XLImage(str(fp))
                w0, h0 = im.width, im.height
                if w0 and h0:
                    scale = min(_THUMB_MAX_PX / float(w0), _THUMB_MAX_PX / float(h0), 1.0)
                    im.width = int(w0 * scale)
                    im.height = int(h0 * scale)
                anchor = f"{get_column_letter(img_i)}{r}"
                ws.add_image(im, anchor)
                if im.height:
                    pts = max(60.0, min(150.0, float(im.height) * 0.72 + 8.0))
                    cur = ws.row_dimensions[r].height
                    ws.row_dimensions[r].height = max(cur or 15.0, pts)
            except Exception:  # noqa: BLE001
                _log.debug("xlsx 嵌入原图失败: %s", pth, exc_info=True)

    for c in range(1, len(cols) + 1):
        name = cols[c - 1]
        if name == XLSX_ORIGINAL_IMAGE_PATH_COLUMN:
            w = 48
        elif name == XLSX_ORIGINAL_IMAGE_EMBED_COLUMN and embed_original_thumbnail:
            w = 22
        elif c <= 2:
            w = 14
        else:
            w = min(50, 18)
        ws.column_dimensions[get_column_letter(c)].width = w


def write_delivery_workbook_to_xlsx(
    path: Path,
    *,
    sheets: Sequence[tuple[Any, ...]],
) -> None:
    """
    写入一个包含多个 sheet 的 xlsx。

    每项为 ``(sheet名, rows, columns)`` 或 ``(sheet名, rows, columns, embed_original_thumbnail)``。
    最后一项为 True 时：在含 ``原图路径`` + ``原图`` 列的 sheet 中，对 ``原图`` 列按路径嵌入缩略图（``合并`` 表应传 False）。
    """
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])
    for spec in sheets:
        if len(spec) == 4:
            title, rows, cols, embed = spec[0], spec[1], spec[2], bool(spec[3])
        elif len(spec) == 3:
            title, rows, cols, embed = spec[0], spec[1], spec[2], False
        else:
            raise TypeError(f"sheet 元组长度须为 3 或 4: {spec!r}")
        _write_sheet(
            wb,
            title=str(title),
            rows=list(rows),
            columns=cols,
            embed_original_thumbnail=embed,
        )
    wb.save(path)


def write_delivery_rows_to_xlsx(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    columns: Sequence[str],
) -> None:
    """兼容旧接口：单 sheet 写入（sheet 名 delivery，不嵌图）。"""
    write_delivery_workbook_to_xlsx(path, sheets=[("delivery", rows, columns, False)])
